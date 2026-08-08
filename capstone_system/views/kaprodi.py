from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.contrib.auth import update_session_auth_hash
import pandas as pd
from django.views.decorators.csrf import csrf_exempt  
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils import timezone


from ..models import (
    AnggotaTim,
    Mahasiswa,
    Dosen,
    DosenCP,
    DosenPembimbing,
    Tim,
    ProposalCapstone,
    Resume,
    PengajuanDospem,
    User,
    Role,
)
from ..forms import DosenForm, ProposalForm, UploadMahasiswaForm
from .base import check_role


# =========================================================
# FUNGSI BANTU: PASTIKAN ROLE ADA DI DATABASE
# =========================================================
def ensure_roles_exist():
    """Pastikan data role selalu ada di database"""
    roles = ['MAHASISWA', 'DOSENCP', 'DOSENPB', 'KAPRODI']
    for role_name in roles:
        Role.objects.get_or_create(name=role_name)


# =========================================================
# DASHBOARD KAPRODI
# =========================================================
@login_required
def kaprodi_home(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    context = {
        'total_proposal': ProposalCapstone.objects.count(),
        'total_resume': Resume.objects.count(),
        'total_mahasiswa': Mahasiswa.objects.filter(status="AKTIF").count(),
        'total_mahasiswa_aktif': Mahasiswa.objects.filter(status="AKTIF").count(),
        'total_mahasiswa_nonaktif': Mahasiswa.objects.filter(status="NONAKTIF").count(),
        'total_dospem': DosenPembimbing.objects.count(),
        'total_dosen': Dosen.objects.filter(status_aktif="AKTIF").count(),
        'total_dosen_aktif': Dosen.objects.filter(status_aktif="AKTIF").count(),
        'total_dosen_nonaktif': Dosen.objects.filter(status_aktif="NONAKTIF").count(),
        'total_tim': Tim.objects.count(),
    }

    return render(request, 'kaprodi/home.html', context)


# =========================================================
# MAHASISWA (MASTER DATA) - DENGAN KATEGORI
# =========================================================
@login_required
def kaprodi_list_mahasiswa(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    # 🔥 Jika tidak ada parameter status, redirect ke ?status=AKTIF
    if not request.GET.get('status') and not request.GET.get('angkatan') and not request.GET.get('q'):
        return redirect(f"{request.path}?status=AKTIF")

    # Ambil parameter filter
    status = request.GET.get('status')
    angkatan = request.GET.get('angkatan')
    keyword = request.GET.get('q', '').strip()
    
    # 🔥 Ambil parameter entries (jumlah data per halaman)
    entries = request.GET.get('entries', '15')  # default 15
    
    # 🔥 Tentukan jumlah per halaman
    if entries == 'all':
        per_page = None  # Semua data
    else:
        try:
            per_page = int(entries)
            if per_page <= 0:
                per_page = 15
        except ValueError:
            per_page = 15

    # Query dasar
    mahasiswa_list = Mahasiswa.objects.select_related('user', 'dosen_pembimbing')

    # Filter status
    if status:
        mahasiswa_list = mahasiswa_list.filter(status=status)
    else:
        mahasiswa_list = mahasiswa_list.filter(status="AKTIF")

    # Filter angkatan
    if angkatan:
        mahasiswa_list = mahasiswa_list.filter(angkatan=angkatan)

    # Filter keyword
    if keyword:
        mahasiswa_list = mahasiswa_list.filter(
            Q(user__first_name__icontains=keyword) |
            Q(user__last_name__icontains=keyword) |
            Q(nim__icontains=keyword)
        )

    mahasiswa_list = mahasiswa_list.order_by('-angkatan', 'nim')

    # Ambil kategori
    for mhs in mahasiswa_list:
        if mhs.kategori:
            mhs.kategori_tim = mhs.kategori
        else:
            anggota = AnggotaTim.objects.filter(
                mahasiswa=mhs,
                status_persetujuan='APPROVED'
            ).first()
            mhs.kategori_tim = anggota.kategori if anggota else None

    daftar_angkatan = Mahasiswa.objects.values_list('angkatan', flat=True).distinct().order_by('-angkatan')

    total_mahasiswa_aktif = Mahasiswa.objects.filter(status="AKTIF").count()
    total_mahasiswa_nonaktif = Mahasiswa.objects.filter(status="NONAKTIF").count()
    total_mahasiswa_arsip = Mahasiswa.objects.filter(status="ARSIP").count()

    # =========================================================
    # 🔥 PAGINATION
    # =========================================================
    if per_page is None:
        # Jika "Semua", tampilkan semua data tanpa pagination
        page_obj = mahasiswa_list
        # Buat paginator palsu untuk template
        class FakePaginator:
            count = mahasiswa_list.count()
            num_pages = 1
            page_range = [1]
        class FakePage:
            def __init__(self, data):
                self.object_list = data
                self.paginator = FakePaginator()
                self.number = 1
                self.has_previous = False
                self.has_next = False
                self.previous_page_number = None
                self.next_page_number = None
                self.start_index = 1
                self.end_index = len(data)
            def __iter__(self):
                return iter(self.object_list)
        page_obj = FakePage(mahasiswa_list)
    else:
        paginator = Paginator(mahasiswa_list, per_page)
        page_number = request.GET.get('page')
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

    context = {
        'page_obj': page_obj,
        'daftar_angkatan': daftar_angkatan,
        'status_filter': status,
        'angkatan_filter': angkatan,
        'keyword': keyword,
        'entries': entries,
        'total_mahasiswa': total_mahasiswa_aktif + total_mahasiswa_nonaktif + total_mahasiswa_arsip,
        'total_mahasiswa_aktif': total_mahasiswa_aktif,
        'total_mahasiswa_nonaktif': total_mahasiswa_nonaktif,
        'total_mahasiswa_arsip': total_mahasiswa_arsip,
    }

    return render(request, 'kaprodi/list_mahasiswa.html', context)


# =========================================================
# NONAKTIFKAN MAHASISWA
# =========================================================
@login_required
@transaction.atomic
def nonaktifkan_mahasiswa(request, id):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    mahasiswa = get_object_or_404(Mahasiswa, id=id)
    
    if mahasiswa.status == "NONAKTIF":
        messages.warning(request, f"{mahasiswa.user.get_full_name()} sudah nonaktif.")
        return redirect("capstone_system:kaprodi_mahasiswa")
    
    if mahasiswa.status == "ARSIP":
        messages.warning(request, f"{mahasiswa.user.get_full_name()} sedang dalam arsip. Kembalikan dulu dari arsip.")
        return redirect("capstone_system:kaprodi_mahasiswa")

    dospem = mahasiswa.dosen_pembimbing

    if mahasiswa.status == "AKTIF" and dospem:
        dospem.update_status()

    mahasiswa.status = "NONAKTIF"
    mahasiswa.save()
    mahasiswa.user.is_active = False
    mahasiswa.user.save()

    messages.success(request, f"{mahasiswa.user.get_full_name()} berhasil dinonaktifkan.")
    return redirect("capstone_system:kaprodi_mahasiswa")


# =========================================================
# AKTIFKAN MAHASISWA
# =========================================================
@login_required
@transaction.atomic
def aktifkan_mahasiswa(request, id):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    mahasiswa = get_object_or_404(Mahasiswa, id=id)
    
    if mahasiswa.status == "AKTIF":
        messages.info(request, f"{mahasiswa.user.get_full_name()} sudah aktif.")
        return redirect("capstone_system:kaprodi_mahasiswa")
    
    if mahasiswa.status == "ARSIP":
        if mahasiswa.dosen_pembimbing:
            dospem = mahasiswa.dosen_pembimbing
            if dospem.penuh:
                messages.error(
                    request, 
                    f"Mahasiswa tidak dapat diaktifkan karena kuota Dosen Pembimbing {dospem.dosen.user.get_full_name()} sudah penuh."
                )
                return redirect("capstone_system:kaprodi_mahasiswa")
            dospem.update_status()
        
        mahasiswa.status = "AKTIF"
        mahasiswa.save()
        mahasiswa.user.is_active = True
        mahasiswa.user.save()
        
        messages.success(request, f"{mahasiswa.user.get_full_name()} berhasil dikembalikan dari arsip.")
        return redirect("capstone_system:kaprodi_mahasiswa")
    
    dospem = mahasiswa.dosen_pembimbing
    if mahasiswa.status == "NONAKTIF" and dospem:
        if dospem.penuh:
            messages.error(
                request, 
                f"Mahasiswa tidak dapat diaktifkan karena kuota Dosen Pembimbing {dospem.dosen.user.get_full_name()} sudah penuh."
            )
            return redirect("capstone_system:kaprodi_mahasiswa")
        dospem.update_status()

    mahasiswa.status = "AKTIF"
    mahasiswa.save()
    mahasiswa.user.is_active = True
    mahasiswa.user.save()

    messages.success(request, f"{mahasiswa.user.get_full_name()} berhasil diaktifkan kembali.")
    return redirect("capstone_system:kaprodi_mahasiswa")


# =========================================================
# KEMBALIKAN MAHASISWA DARI ARSIP
# =========================================================
@login_required
@transaction.atomic
def kembalikan_dari_arsip(request, id):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    mahasiswa = get_object_or_404(Mahasiswa, id=id)
    
    if mahasiswa.status != "ARSIP":
        messages.warning(request, f"Mahasiswa {mahasiswa.user.get_full_name()} tidak dalam status arsip.")
        return redirect("capstone_system:kaprodi_mahasiswa")

    if mahasiswa.dosen_pembimbing:
        dospem = mahasiswa.dosen_pembimbing
        if dospem.penuh:
            messages.error(
                request, 
                f"Mahasiswa tidak dapat dikembalikan karena kuota Dosen Pembimbing {dospem.dosen.user.get_full_name()} sudah penuh."
            )
            return redirect("capstone_system:arsip_mahasiswa")
        dospem.update_status()

    mahasiswa.status = "AKTIF"
    mahasiswa.save()
    mahasiswa.user.is_active = True
    mahasiswa.user.save()

    messages.success(
        request, 
        f"{mahasiswa.user.get_full_name()} berhasil dikembalikan dari arsip."
    )
    return redirect("capstone_system:arsip_mahasiswa")


# =========================================================
# DETAIL MAHASISWA
# =========================================================
@login_required
def detail_mahasiswa(request, id):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    mahasiswa = get_object_or_404(Mahasiswa, id=id)
    
    if mahasiswa.kategori:
        mahasiswa.kategori_tim = mahasiswa.kategori
    else:
        anggota = AnggotaTim.objects.filter(
            mahasiswa=mahasiswa,
            status_persetujuan='APPROVED'
        ).first()
        mahasiswa.kategori_tim = anggota.kategori if anggota else None
    
    tim_ids = AnggotaTim.objects.filter(
        mahasiswa=mahasiswa,
        status_persetujuan='APPROVED'
    ).values_list('tim_id', flat=True)
    
    proposal_list = ProposalCapstone.objects.filter(
        tim_id__in=tim_ids
    ).order_by('-waktu_pengajuan')
    
    resume_list = Resume.objects.filter(
        mahasiswa=mahasiswa
    ).order_by('-waktu_pengajuan')
    
    pengajuan_list = PengajuanDospem.objects.filter(
        mahasiswa=mahasiswa
    ).order_by('-tanggal_pengajuan').select_related(
        'dosen_pembimbing__dosen__user'
    )

    return render(request, "kaprodi/detail_mahasiswa.html", {
        "mahasiswa": mahasiswa,
        "proposal_list": proposal_list,
        "resume_list": resume_list,
        "pengajuan_list": pengajuan_list,
    })


# =========================================================
# TAMBAH MAHASISWA
# =========================================================
@login_required
@transaction.atomic
def tambah_mahasiswa(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    # Pastikan role MAHASISWA ada
    ensure_roles_exist()

    if request.method == "POST":
        nim = request.POST.get('nim', '').strip()
        nama = request.POST.get('nama_lengkap', '').strip()
        email = request.POST.get('email', '').strip()
        kelas = request.POST.get('kelas', '').strip().upper()
        angkatan = request.POST.get('angkatan', '').strip()
        tanggal_masuk = request.POST.get('tanggal_masuk', '').strip()

        if not nim:
            messages.error(request, "NIM wajib diisi!")
            return redirect('capstone_system:tambah_mahasiswa')
        
        if not nama:
            messages.error(request, "Nama lengkap wajib diisi!")
            return redirect('capstone_system:tambah_mahasiswa')
        
        if not email:
            messages.error(request, "Email wajib diisi!")
            return redirect('capstone_system:tambah_mahasiswa')
        
        if not kelas:
            messages.error(request, "Kelas wajib diisi!")
            return redirect('capstone_system:tambah_mahasiswa')
        
        if not angkatan:
            messages.error(request, "Angkatan wajib diisi!")
            return redirect('capstone_system:tambah_mahasiswa')

        if User.objects.filter(username=nim).exists():
            messages.error(request, f"NIM '{nim}' sudah terdaftar sebagai User!")
            return redirect('capstone_system:tambah_mahasiswa')

        if Mahasiswa.objects.filter(nim=nim).exists():
            messages.error(request, f"NIM '{nim}' sudah terdaftar sebagai Mahasiswa!")
            return redirect('capstone_system:tambah_mahasiswa')

        if User.objects.filter(email=email).exists():
            messages.error(request, f"Email '{email}' sudah digunakan!")
            return redirect('capstone_system:tambah_mahasiswa')

        try:
            def split_name(full_name):
                if not full_name:
                    return '', ''
                full_name = full_name.strip()
                parts = full_name.split()
                if len(parts) > 1:
                    return parts[0], ' '.join(parts[1:])
                return full_name, ''

            first_name, last_name = split_name(nama)

            user = User.objects.create_user(
                username=nim,
                password=nim,
                first_name=first_name,
                last_name=last_name,
                email=email,
                role='MAHASISWA',
                is_password_changed=False,
                is_active=True,
            )

            role_mahasiswa, created = Role.objects.get_or_create(name='MAHASISWA')
            user.roles.add(role_mahasiswa)

            if tanggal_masuk:
                try:
                    from datetime import datetime
                    tanggal_masuk_obj = datetime.strptime(tanggal_masuk, '%Y-%m-%d').date()
                except ValueError:
                    tanggal_masuk_obj = timezone.now().date()
            else:
                tanggal_masuk_obj = timezone.now().date()

            mahasiswa = Mahasiswa.objects.create(
                user=user,
                nim=nim,
                kelas=kelas,
                angkatan=angkatan,
                status='AKTIF',
                tanggal_masuk=tanggal_masuk_obj,
                kategori=None,
                dosen_pembimbing=None,
                foto_profil=None,
            )

            messages.success(request, f"✅ Mahasiswa {nama} (NIM: {nim}) berhasil ditambahkan!")
            return redirect("capstone_system:kaprodi_mahasiswa")

        except Exception as e:
            messages.error(request, f"❌ Gagal menyimpan data: {str(e)}")
            return redirect('capstone_system:tambah_mahasiswa')

    return render(request, "kaprodi/tambah_mahasiswa.html")


# =========================================================
# EDIT MAHASISWA
# =========================================================
@login_required
def edit_mahasiswa(request, id):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    mahasiswa = get_object_or_404(Mahasiswa, id=id)
    user = mahasiswa.user

    dospem_list = DosenPembimbing.objects.filter(
        Q(status="OPEN") | Q(id=mahasiswa.dosen_pembimbing_id)
    ).select_related("dosen__user")

    next_url = request.GET.get('next') or request.POST.get('next')
    
    anggota = AnggotaTim.objects.filter(
        mahasiswa=mahasiswa,
        status_persetujuan='APPROVED'
    ).first()
    mahasiswa.kategori_tim = anggota.kategori if anggota else None

    if request.method == "POST":
        user.first_name = request.POST.get("nama")
        user.email = request.POST.get("email")
        nim = request.POST.get("nim")

        if Mahasiswa.objects.exclude(id=mahasiswa.id).filter(nim=nim).exists():
            messages.error(request, "NIM sudah digunakan.")
            return redirect("capstone_system:edit_mahasiswa", id=id)

        user.username = nim
        user.save()

        mahasiswa.nim = nim
        mahasiswa.kelas = request.POST.get("kelas")
        mahasiswa.angkatan = request.POST.get("angkatan")
        mahasiswa.kategori = request.POST.get("kategori")
        mahasiswa.status = request.POST.get("status")
        mahasiswa.tanggal_masuk = request.POST.get("tanggal_masuk") or None

        dospem_id = request.POST.get("dosen_pembimbing")
        dospem_lama = mahasiswa.dosen_pembimbing

        if not dospem_id:
            if dospem_lama:
                dospem_lama.update_status()
            mahasiswa.dosen_pembimbing = None
        else:
            dospem_baru = get_object_or_404(DosenPembimbing, id=dospem_id)
            if dospem_lama is None or dospem_lama.id != dospem_baru.id:
                if dospem_baru.penuh:
                    messages.error(request, f"Kuota Dosen Pembimbing {dospem_baru.dosen.user.get_full_name()} sudah penuh.")
                    return redirect("capstone_system:edit_mahasiswa", id=id)
                if dospem_lama:
                    dospem_lama.update_status()
                mahasiswa.dosen_pembimbing = dospem_baru
                dospem_baru.update_status()
            else:
                mahasiswa.dosen_pembimbing = dospem_lama

        mahasiswa.save()
        if mahasiswa.dosen_pembimbing:
            mahasiswa.dosen_pembimbing.update_status()

        messages.success(request, "Data mahasiswa berhasil diperbarui.")

        if next_url:
            return redirect(next_url)
        return redirect("capstone_system:kaprodi_mahasiswa")

    return render(request, "kaprodi/edit_mahasiswa.html", {
        "mahasiswa": mahasiswa,
        "dospem_list": dospem_list,
        "next_url": next_url,
    })


# =========================================================
# DOSEN PEMBIMBING (MANAGE)
# =========================================================
@login_required
def kaprodi_manage_dospem(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    # Pastikan role ada
    ensure_roles_exist()

    keyword = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    entries = request.GET.get('entries', '10')

    dospem_list = DosenPembimbing.objects.select_related(
        "dosen", "dosen__user"
    ).filter(dosen__status_aktif="AKTIF").order_by("dosen__user__first_name")

    if keyword:
        dospem_list = dospem_list.filter(
            Q(dosen__user__first_name__icontains=keyword) |
            Q(dosen__user__last_name__icontains=keyword)
        )

    if status_filter:
        dospem_list = dospem_list.filter(status=status_filter)

    for dospem in dospem_list:
        dospem.update_status()

    if entries == 'all':
        per_page = None
    else:
        try:
            per_page = int(entries)
            if per_page <= 0:
                per_page = 10
        except ValueError:
            per_page = 10

    if per_page is None:
        class FakePaginator:
            count = dospem_list.count()
            num_pages = 1
            page_range = [1]
        class FakePage:
            def __init__(self, data):
                self.object_list = data
                self.paginator = FakePaginator()
                self.number = 1
                self.has_previous = False
                self.has_next = False
                self.previous_page_number = None
                self.next_page_number = None
                self.start_index = 1
                self.end_index = len(data)
            def __iter__(self):
                return iter(self.object_list)
        page_obj = FakePage(dospem_list)
    else:
        paginator = Paginator(dospem_list, per_page)
        page_number = request.GET.get('page')
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

    context = {
        "page_obj": page_obj,
        "keyword": keyword,
        "status_filter": status_filter,
        "entries": entries,
        "total_dospem": dospem_list.count(),
        "total_open": dospem_list.filter(status="OPEN").count(),
        "total_full": dospem_list.filter(status="FULL").count(),
    }

    return render(request, "kaprodi/manage_dospem.html", context)


# =========================================================
# TAMBAH DOSEN (DENGAN MULTI-ROLE)
# =========================================================
@login_required
@transaction.atomic
def tambah_dosen(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    # 🔥 PASTIKAN ROLE ADA DI DATABASE
    ensure_roles_exist()

    if request.method == "POST":
        form = DosenForm(request.POST)
        if form.is_valid():
            dosen = form.save()
            messages.success(request, f"Dosen {dosen.user.get_full_name()} berhasil ditambahkan.")
            return redirect("capstone_system:kaprodi_dosen")
        else:
            messages.error(request, "Terjadi kesalahan. Silakan periksa kembali form.")
    else:
        form = DosenForm()

    return render(request, "kaprodi/tambah_dosen.html", {"form": form})


# =========================================================
# EDIT DOSEN (DENGAN MULTI-ROLE)
# =========================================================
@login_required
@transaction.atomic
def edit_dosen(request, id):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    # 🔥 PASTIKAN ROLE ADA DI DATABASE
    ensure_roles_exist()

    dosen = get_object_or_404(Dosen, id=id)

    if request.method == "POST":
        form = DosenForm(request.POST, instance=dosen)
        if form.is_valid():
            dosen = form.save()
            messages.success(request, "Data dosen berhasil diperbarui.")
            return redirect("capstone_system:kaprodi_dosen")
        else:
            messages.error(request, "Terjadi kesalahan. Silakan periksa kembali form.")
    else:
        form = DosenForm(instance=dosen)

    return render(request, "kaprodi/edit_dosen.html", {
        "form": form,
        "dosen": dosen
    })


# =========================================================
# LIST DOSEN
# =========================================================
@login_required
def kaprodi_list_dosen(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    keyword = request.GET.get('q', '')
    status = request.GET.get('status')
    entries = request.GET.get('entries', '10')

    dosen_list = Dosen.objects.select_related('user').all()

    if keyword:
        dosen_list = dosen_list.filter(
            Q(user__first_name__icontains=keyword) |
            Q(user__last_name__icontains=keyword) |
            Q(nip__icontains=keyword)
        )

    if status:
        dosen_list = dosen_list.filter(status_aktif=status)

    dosen_list = dosen_list.order_by('user__first_name')

    if entries == 'all':
        per_page = None
    else:
        try:
            per_page = int(entries)
            if per_page <= 0:
                per_page = 10
        except ValueError:
            per_page = 10

    if per_page is None:
        class FakePaginator:
            count = dosen_list.count()
            num_pages = 1
            page_range = [1]
        class FakePage:
            def __init__(self, data):
                self.object_list = data
                self.paginator = FakePaginator()
                self.number = 1
                self.has_previous = False
                self.has_next = False
                self.previous_page_number = None
                self.next_page_number = None
                self.start_index = 1
                self.end_index = len(data)
            def __iter__(self):
                return iter(self.object_list)
        page_obj = FakePage(dosen_list)
    else:
        paginator = Paginator(dosen_list, per_page)
        page_number = request.GET.get('page')
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

    context = {
        'page_obj': page_obj,
        'keyword': keyword,
        'status_filter': status,
        'entries': entries,
        'total_dosen': Dosen.objects.filter(status_aktif='AKTIF').count(),
        'total_dosen_aktif': Dosen.objects.filter(status_aktif='AKTIF').count(),
        'total_dosen_nonaktif': Dosen.objects.filter(status_aktif='NONAKTIF').count(),
    }

    return render(request, 'kaprodi/list_dosen.html', context)


# =========================================================
# NON AKTIF DOSEN
# =========================================================
@login_required
def nonaktifkan_dosen(request, id):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    dosen = get_object_or_404(Dosen, id=id)
    dosen.status_aktif = "NONAKTIF"
    dosen.save()
    dosen.user.is_active = False
    dosen.user.save()
    DosenPembimbing.objects.filter(dosen=dosen).update(status="CLOSED")

    messages.success(request, "Dosen berhasil dinonaktifkan.")
    return redirect('capstone_system:kaprodi_dosen')


# =========================================================
# AKTIF DOSEN
# =========================================================
@login_required
def aktifkan_dosen(request, id):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    dosen = get_object_or_404(Dosen, id=id)
    dosen.status_aktif = "AKTIF"
    dosen.save()
    dosen.user.is_active = True
    dosen.user.save()
    DosenPembimbing.objects.filter(dosen=dosen).update(status="OPEN")

    messages.success(request, "Dosen berhasil diaktifkan.")
    return redirect('capstone_system:kaprodi_dosen')


# =========================================================
# HAPUS DOSEN
# =========================================================
@login_required
@transaction.atomic
def hapus_dosen(request, id):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    dosen = get_object_or_404(Dosen, id=id)

    if Mahasiswa.objects.filter(dosen_pembimbing__dosen=dosen).exists():
        messages.error(request, "Dosen tidak dapat dihapus karena masih menjadi pembimbing mahasiswa.")
        return redirect("capstone_system:kaprodi_dosen")

    DosenPembimbing.objects.filter(dosen=dosen).delete()
    DosenCP.objects.filter(dosen=dosen).delete()
    dosen.user.delete()

    messages.success(request, "Data dosen berhasil dihapus.")
    return redirect("capstone_system:kaprodi_dosen")


# =========================================================
# TIM
# =========================================================
@login_required
def kaprodi_list_tim(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    keyword = request.GET.get('q', '')
    kategori_filter = request.GET.get('kategori', '')
    status_filter = request.GET.get('status', '')
    entries = request.GET.get('entries', '10')

    tim_list = Tim.objects.all().order_by('-dibuat_pada')

    if keyword:
        tim_list = tim_list.filter(nama_tim__icontains=keyword)

    if kategori_filter:
        tim_list = tim_list.filter(kategori=kategori_filter)

    if status_filter:
        tim_list = tim_list.filter(status=status_filter)

    if entries == 'all':
        per_page = None
    else:
        try:
            per_page = int(entries)
            if per_page <= 0:
                per_page = 10
        except ValueError:
            per_page = 10

    if per_page is None:
        class FakePaginator:
            count = tim_list.count()
            num_pages = 1
            page_range = [1]
        class FakePage:
            def __init__(self, data):
                self.object_list = data
                self.paginator = FakePaginator()
                self.number = 1
                self.has_previous = False
                self.has_next = False
                self.previous_page_number = None
                self.next_page_number = None
                self.start_index = 1
                self.end_index = len(data)
            def __iter__(self):
                return iter(self.object_list)
        page_obj = FakePage(tim_list)
    else:
        paginator = Paginator(tim_list, per_page)
        page_number = request.GET.get('page')
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

    context = {
        'page_obj': page_obj,
        'keyword': keyword,
        'kategori_filter': kategori_filter,
        'status_filter': status_filter,
        'entries': entries,
    }

    return render(request, 'kaprodi/list_tim.html', context)


@login_required
def kaprodi_detail_tim(request, tim_id):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    tim = get_object_or_404(Tim, id=tim_id)
    anggota = tim.anggota.all()
    return render(request, 'kaprodi/detail_tim.html', {
        'tim': tim,
        'anggota': anggota
    })


# =========================================================
# PROPOSAL
# =========================================================
@login_required
def kaprodi_list_proposal(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    proposals = ProposalCapstone.objects.all().order_by('-waktu_pengajuan')
    return render(request, 'kaprodi/list_proposal.html', {'proposals': proposals})


@login_required
def kaprodi_detail_proposal(request, proposal_id):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    proposal = get_object_or_404(ProposalCapstone, id=proposal_id)
    return render(request, 'kaprodi/detail_proposal.html', {'proposal': proposal})


@login_required
def kaprodi_update_proposal(request, proposal_id):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    proposal = get_object_or_404(ProposalCapstone, id=proposal_id)

    if request.method == 'POST':
        form = ProposalForm(request.POST, request.FILES, instance=proposal)
        if form.is_valid():
            form.save()
            messages.success(request, "Proposal berhasil diperbarui")
            return redirect('capstone_system:kaprodi_proposal')
    else:
        form = ProposalForm(instance=proposal)

    return render(request, 'kaprodi/update_proposal.html', {
        'form': form,
        'proposal': proposal
    })


@login_required
def kaprodi_delete_proposal(request, proposal_id):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    proposal = get_object_or_404(ProposalCapstone, id=proposal_id)

    if request.method == 'POST':
        proposal.delete()
        messages.success(request, "Proposal berhasil dihapus")
        return redirect('capstone_system:kaprodi_proposal')

    return render(request, 'kaprodi/delete_proposal.html', {'proposal': proposal})


# =========================================================
# RESUME
# =========================================================
@login_required
def kaprodi_list_resume(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    resumes = Resume.objects.all().order_by('-waktu_pengajuan')
    return render(request, 'kaprodi/list_resume.html', {'resumes': resumes})


@login_required
def kaprodi_detail_resume(request, resume_id):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    resume = get_object_or_404(Resume, id=resume_id)
    return render(request, 'kaprodi/detail_resume.html', {'resume': resume})


# =========================================================
# PENGAJUAN DOSPEM
# =========================================================
@login_required
def kaprodi_pengajuan(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    keyword = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    dosen_filter = request.GET.get('dosen', '')
    entries = request.GET.get('entries', '10')

    pengajuan_list = PengajuanDospem.objects.all().order_by('-tanggal_pengajuan').select_related(
        'mahasiswa__user', 'dosen_pembimbing__dosen__user'
    )

    if keyword:
        pengajuan_list = pengajuan_list.filter(
            Q(mahasiswa__user__first_name__icontains=keyword) |
            Q(mahasiswa__user__last_name__icontains=keyword) |
            Q(mahasiswa__nim__icontains=keyword)
        )

    if status_filter:
        pengajuan_list = pengajuan_list.filter(status=status_filter)

    if dosen_filter:
        pengajuan_list = pengajuan_list.filter(dosen_pembimbing_id=dosen_filter)

    dosen_list = DosenPembimbing.objects.select_related('dosen__user').all().order_by('dosen__user__first_name')

    if entries == 'all':
        per_page = None
    else:
        try:
            per_page = int(entries)
            if per_page <= 0:
                per_page = 10
        except ValueError:
            per_page = 10

    if per_page is None:
        class FakePaginator:
            count = pengajuan_list.count()
            num_pages = 1
            page_range = [1]
        class FakePage:
            def __init__(self, data):
                self.object_list = data
                self.paginator = FakePaginator()
                self.number = 1
                self.has_previous = False
                self.has_next = False
                self.previous_page_number = None
                self.next_page_number = None
                self.start_index = 1
                self.end_index = len(data)
            def __iter__(self):
                return iter(self.object_list)
        page_obj = FakePage(pengajuan_list)
    else:
        paginator = Paginator(pengajuan_list, per_page)
        page_number = request.GET.get('page')
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

    context = {
        'page_obj': page_obj,
        'keyword': keyword,
        'status_filter': status_filter,
        'dosen_filter': dosen_filter,
        'dosen_list': dosen_list,
        'entries': entries,
    }

    return render(request, 'kaprodi/pengajuan.html', context)


# =========================================================
# MONITORING CAPSTONE
# =========================================================
@login_required
def kaprodi_monitoring(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    keyword = request.GET.get('q', '').strip()
    status_proposal = request.GET.get('status_proposal', '')
    status_resume = request.GET.get('status_resume', '')
    sort_by = request.GET.get('sort', 'nama')
    sort_order = request.GET.get('order', 'asc')
    entries = request.GET.get('entries', '10')

    mahasiswa_list = Mahasiswa.objects.filter(status='AKTIF').select_related(
        'user', 'dosen_pembimbing', 'dosen_pembimbing__dosen', 'dosen_pembimbing__dosen__user'
    )

    if keyword:
        mahasiswa_list = mahasiswa_list.filter(
            Q(user__first_name__icontains=keyword) |
            Q(user__last_name__icontains=keyword) |
            Q(nim__icontains=keyword)
        )

    monitoring_data = []
    total_proposal = 0
    total_resume = 0
    total_belum_upload = 0

    stats_proposal = {
        'diterima': 0,
        'ditolak': 0,
        'revisi': 0,
        'pending': 0,
        'belum_review': 0,
    }
    stats_resume = {
        'diterima': 0,
        'ditolak': 0,
        'revisi': 0,
        'pending': 0,
        'belum_review': 0,
    }

    semua_dospem = DosenPembimbing.objects.filter(
        dosen__status_aktif='AKTIF'
    ).select_related('dosen', 'dosen__user')

    dosen_dict = {}
    for dospem in semua_dospem:
        nama_dosen = dospem.dosen.user.get_full_name()
        dosen_dict[nama_dosen] = {
            'jumlah_mahasiswa': 0,
            'dospem_id': dospem.id,
            'status': dospem.status,
            'batas_bimbingan': dospem.batas_bimbingan,
        }

    total_tanpa_dosen = 0

    for mhs in mahasiswa_list:
        proposal = ProposalCapstone.objects.filter(
            tim__anggota__mahasiswa=mhs,
            tim__anggota__status_persetujuan='APPROVED'
        ).order_by('-waktu_pengajuan').first()

        resume = Resume.objects.filter(
            mahasiswa=mhs
        ).order_by('-waktu_pengajuan').first()

        status_proposal_text = None
        if proposal:
            status_proposal_text = proposal.status_final
            total_proposal += 1
            
            if status_proposal_text == 'DITERIMA':
                stats_proposal['diterima'] += 1
            elif status_proposal_text == 'DITOLAK':
                stats_proposal['ditolak'] += 1
            elif status_proposal_text == 'REVISI':
                stats_proposal['revisi'] += 1
            elif status_proposal_text == 'SEDANG_REVIEW':
                stats_proposal['pending'] += 1
            elif status_proposal_text == 'BELUM_REVIEW':
                stats_proposal['belum_review'] += 1
        else:
            status_proposal_text = 'BELUM_UPLOAD'

        status_resume_text = None
        if resume:
            status_resume_text = resume.status
            total_resume += 1
            
            if status_resume_text == 'DISETUJUI':
                stats_resume['diterima'] += 1
            elif status_resume_text == 'DITOLAK':
                stats_resume['ditolak'] += 1
            elif status_resume_text == 'REVISI':
                stats_resume['revisi'] += 1
            elif status_resume_text == 'BELUM_REVIEW':
                stats_resume['belum_review'] += 1
                stats_resume['pending'] += 1
        else:
            status_resume_text = 'BELUM_UPLOAD'

        if status_proposal and status_proposal_text != status_proposal:
            continue
        if status_resume and status_resume_text != status_resume:
            continue

        if not proposal and not resume:
            total_belum_upload += 1

        dosen_nama = None
        if mhs.dosen_pembimbing:
            dosen_nama = mhs.dosen_pembimbing.dosen.user.get_full_name()
            if dosen_nama in dosen_dict:
                dosen_dict[dosen_nama]['jumlah_mahasiswa'] += 1
            else:
                dosen_dict[dosen_nama] = {
                    'jumlah_mahasiswa': 1,
                    'dospem_id': mhs.dosen_pembimbing.id,
                    'status': mhs.dosen_pembimbing.status,
                    'batas_bimbingan': mhs.dosen_pembimbing.batas_bimbingan,
                }
        else:
            total_tanpa_dosen += 1

        monitoring_data.append({
            'mahasiswa_id': mhs.id,
            'nim': mhs.nim,
            'nama': mhs.user.get_full_name(),
            'kelas': mhs.kelas,
            'angkatan': mhs.angkatan,
            'status_proposal': status_proposal_text,
            'status_resume': status_resume_text,
            'proposal_id': proposal.id if proposal else None,
            'resume_id': resume.id if resume else None,
            'dosen_pembimbing': dosen_nama,
        })

    sort_mapping = {
        'nama': 'nama',
        'nim': 'nim',
        'kelas': 'kelas',
        'angkatan': 'angkatan',
        'status_proposal': 'status_proposal',
        'status_resume': 'status_resume',
    }

    sort_field = sort_mapping.get(sort_by, 'nama')
    reverse = sort_order == 'desc'
    monitoring_data.sort(key=lambda x: x.get(sort_field, '').lower() if x.get(sort_field) else '', reverse=reverse)

    dosen_list = []
    for nama, data in dosen_dict.items():
        dosen_list.append({
            'nama': nama,
            'jumlah_mahasiswa': data['jumlah_mahasiswa'],
            'dospem_id': data['dospem_id'],
            'status': data['status'],
            'batas_bimbingan': data['batas_bimbingan'],
            'sisa_kuota': data['batas_bimbingan'] - data['jumlah_mahasiswa'],
            'penuh': data['jumlah_mahasiswa'] >= data['batas_bimbingan'],
        })
    
    dosen_list.sort(key=lambda x: x['jumlah_mahasiswa'], reverse=True)

    if entries == 'all':
        per_page = None
    else:
        try:
            per_page = int(entries)
            if per_page <= 0:
                per_page = 10
        except ValueError:
            per_page = 10

    if per_page is None:
        class FakePaginator:
            count = len(monitoring_data)
            num_pages = 1
            page_range = [1]
        class FakePage:
            def __init__(self, data):
                self.object_list = data
                self.paginator = FakePaginator()
                self.number = 1
                self.has_previous = False
                self.has_next = False
                self.previous_page_number = None
                self.next_page_number = None
                self.start_index = 1
                self.end_index = len(data)
            def __iter__(self):
                return iter(self.object_list)
        page_obj = FakePage(monitoring_data)
    else:
        paginator = Paginator(monitoring_data, per_page)
        page_number = request.GET.get('page')
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

    context = {
        'page_obj': page_obj,
        'total_mahasiswa': mahasiswa_list.count(),
        'total_proposal': total_proposal,
        'total_resume': total_resume,
        'total_belum_upload': total_belum_upload,
        'stats_proposal': stats_proposal,
        'stats_resume': stats_resume,
        'keyword': keyword,
        'status_proposal': status_proposal,
        'status_resume': status_resume,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'entries': entries,
        'dosen_list': dosen_list,
        'total_tanpa_dosen': total_tanpa_dosen,
    }

    return render(request, 'kaprodi/monitoring.html', context)


# =========================================================
# LAPORAN
# =========================================================
@login_required
def kaprodi_laporan(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    total_mahasiswa = Mahasiswa.objects.count()
    total_dosen = Dosen.objects.count()
    total_dospem = DosenPembimbing.objects.count()
    total_dosen_cp = DosenCP.objects.count()
    total_tim = Tim.objects.count()
    total_proposal = ProposalCapstone.objects.count()
    total_resume = Resume.objects.count()

    return render(request, 'kaprodi/laporan.html', {
        'total_mahasiswa': total_mahasiswa,
        'total_dosen': total_dosen,
        'total_dospem': total_dospem,
        'total_dosen_cp': total_dosen_cp,
        'total_tim': total_tim,
        'total_proposal': total_proposal,
        'total_resume': total_resume,
    })


# =========================================================
# PROFILE KAPRODI
# =========================================================
@login_required
def kaprodi_profile(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    user = request.user
    dosen, created = Dosen.objects.get_or_create(user=user)

    if request.method == 'POST':
        user.email = request.POST.get('email', '').strip()
        password = request.POST.get('password')
        if password:
            user.set_password(password)
            update_session_auth_hash(request, user)
        user.save()

        if request.FILES.get('foto_profil'):
            dosen.foto_profil = request.FILES['foto_profil']
            dosen.save()

        messages.success(request, "Profil berhasil diperbarui.")
        return redirect('capstone_system:kaprodi_profile')

    return render(request, 'kaprodi/profile.html', {
        'user': user,
        'dosen': dosen,
    })


# =========================================================
# ARSIP MAHASISWA
# =========================================================
@login_required
def arsip_mahasiswa(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    keyword = request.GET.get("q", "")
    entries = request.GET.get('entries', '10')

    if entries == 'all':
        per_page = None
    else:
        try:
            per_page = int(entries)
            if per_page <= 0:
                per_page = 10
        except ValueError:
            per_page = 10

    mahasiswa_list = Mahasiswa.objects.filter(status="ARSIP").select_related("user", "dosen_pembimbing")

    if keyword:
        mahasiswa_list = mahasiswa_list.filter(
            Q(user__first_name__icontains=keyword) |
            Q(user__last_name__icontains=keyword) |
            Q(nim__icontains=keyword)
        )

    mahasiswa_list = mahasiswa_list.order_by('-angkatan', 'nim')

    if per_page is None:
        class FakePaginator:
            count = mahasiswa_list.count()
            num_pages = 1
            page_range = [1]
        class FakePage:
            def __init__(self, data):
                self.object_list = data
                self.paginator = FakePaginator()
                self.number = 1
                self.has_previous = False
                self.has_next = False
                self.previous_page_number = None
                self.next_page_number = None
                self.start_index = 1
                self.end_index = len(data)
            def __iter__(self):
                return iter(self.object_list)
        page_obj = FakePage(mahasiswa_list)
    else:
        paginator = Paginator(mahasiswa_list, per_page)
        page_number = request.GET.get('page')
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

    context = {
        "page_obj": page_obj,
        "keyword": keyword,
        "entries": entries,
        "total_arsip": mahasiswa_list.count(),
    }

    return render(request, "kaprodi/arsip_mahasiswa.html", context)


# =========================================================
# LIST DOSEN CAPSTONE
# =========================================================
@login_required
def kaprodi_dosen_cp(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    keyword = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    entries = request.GET.get('entries', '10')

    dosen_capstone = DosenCP.objects.select_related(
        "dosen", "dosen__user"
    ).all().order_by("dosen__user__first_name")

    if keyword:
        dosen_capstone = dosen_capstone.filter(
            Q(dosen__user__first_name__icontains=keyword) |
            Q(dosen__user__last_name__icontains=keyword) |
            Q(dosen__nip__icontains=keyword)
        )

    if status_filter:
        dosen_capstone = dosen_capstone.filter(dosen__status_aktif=status_filter)

    if entries == 'all':
        per_page = None
    else:
        try:
            per_page = int(entries)
            if per_page <= 0:
                per_page = 10
        except ValueError:
            per_page = 10

    if per_page is None:
        class FakePaginator:
            count = dosen_capstone.count()
            num_pages = 1
            page_range = [1]
        class FakePage:
            def __init__(self, data):
                self.object_list = data
                self.paginator = FakePaginator()
                self.number = 1
                self.has_previous = False
                self.has_next = False
                self.previous_page_number = None
                self.next_page_number = None
                self.start_index = 1
                self.end_index = len(data)
            def __iter__(self):
                return iter(self.object_list)
        page_obj = FakePage(dosen_capstone)
    else:
        paginator = Paginator(dosen_capstone, per_page)
        page_number = request.GET.get('page')
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

    context = {
        "page_obj": page_obj,
        "keyword": keyword,
        "status_filter": status_filter,
        "entries": entries,
        "total_dosen_capstone": dosen_capstone.count(),
    }

    return render(request, "kaprodi/list_dosen_cp.html", context)


# =========================================================
# ARSIP TAHUNAN MAHASISWA
# =========================================================
@login_required
@transaction.atomic
def arsip_tahunan(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    if request.method == 'POST':
        ids = request.POST.getlist('mahasiswa')
        if not ids:
            messages.warning(request, 'Pilih minimal satu mahasiswa.')
            return redirect('capstone_system:kaprodi_mahasiswa')

        mahasiswa_list = Mahasiswa.objects.filter(id__in=ids).select_related('user', 'dosen_pembimbing')
        
        dospem_to_update = set()

        for mahasiswa in mahasiswa_list:
            if mahasiswa.status in ["AKTIF", "NONAKTIF"]:
                if mahasiswa.status == "AKTIF" and mahasiswa.dosen_pembimbing:
                    dospem_to_update.add(mahasiswa.dosen_pembimbing)

                mahasiswa.status = 'ARSIP'
                mahasiswa.save()
                mahasiswa.user.is_active = False
                mahasiswa.user.save()

        for dospem in dospem_to_update:
            dospem.update_status()

        messages.success(request, f'{len(mahasiswa_list)} mahasiswa berhasil diarsipkan.')

    return redirect('capstone_system:kaprodi_mahasiswa')


# =========================================================
# EDIT BATAS BIMBINGAN DOSEN PEMBIMBING
# =========================================================
@login_required
def edit_batas_dospem(request, id):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    dospem = get_object_or_404(DosenPembimbing, id=id)

    if request.method == "POST":
        batas = int(request.POST.get("batas_bimbingan"))
        if batas < dospem.jumlah_bimbingan:
            messages.error(request, f"Batas minimal adalah {dospem.jumlah_bimbingan} karena saat ini dosen masih membimbing {dospem.jumlah_bimbingan} mahasiswa.")
            return redirect("capstone_system:edit_batas_dospem", id=id)

        dospem.batas_bimbingan = batas
        dospem.save()
        dospem.update_status()

        messages.success(request, "Batas bimbingan berhasil diperbarui.")
        return redirect("capstone_system:kaprodi_dospem")

    return render(request, "kaprodi/edit_batas_dospem.html", {"dospem": dospem})


# =========================================================
# UPLOAD MAHASISWA VIA EXCEL
# =========================================================
@login_required
def upload_mahasiswa_view(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response
    
    form = UploadMahasiswaForm()
    context = {
        'form': form,
        'title': 'Upload Data Mahasiswa',
        'breadcrumb': [
            {'name': 'Dashboard', 'url': 'capstone_system:kaprodi_home'},
            {'name': 'Upload Mahasiswa', 'url': '#'},
        ]
    }
    return render(request, 'kaprodi/upload_mahasiswa.html', context)


@login_required
@csrf_exempt
def upload_mahasiswa_process(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method tidak diizinkan'}, status=405)
    
    form = UploadMahasiswaForm(request.POST, request.FILES)
    
    if not form.is_valid():
        return JsonResponse({
            'success': False,
            'message': 'Validasi gagal',
            'errors': form.errors
        }, status=400)
    
    file = request.FILES['file']
    
    try:
        df = pd.read_excel(file)
        
        required_columns = ['nim', 'nama', 'kelas', 'angkatan']
        df.columns = df.columns.str.lower().str.strip()
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return JsonResponse({
                'success': False,
                'message': f'Kolom yang diperlukan tidak ditemukan: {", ".join(missing_columns)}'
            }, status=400)
        
        imported_data = []
        errors = []
        success_count = 0
        
        with transaction.atomic():
            for index, row in df.iterrows():
                row_num = index + 2
                
                nim = str(row['nim']).strip()
                nama = str(row['nama']).strip()
                kelas = str(row['kelas']).strip()
                angkatan = str(row['angkatan']).strip()
                
                if not nim:
                    errors.append(f"Baris {row_num}: NIM tidak boleh kosong")
                    continue
                
                if not nama:
                    errors.append(f"Baris {row_num}: Nama tidak boleh kosong")
                    continue
                
                if not kelas:
                    errors.append(f"Baris {row_num}: Kelas tidak boleh kosong")
                    continue
                
                if not angkatan:
                    errors.append(f"Baris {row_num}: Angkatan tidak boleh kosong")
                    continue
                
                if User.objects.filter(username=nim).exists():
                    errors.append(f"Baris {row_num}: NIM '{nim}' sudah terdaftar sebagai User")
                    continue
                
                if Mahasiswa.objects.filter(nim=nim).exists():
                    errors.append(f"Baris {row_num}: NIM '{nim}' sudah terdaftar sebagai Mahasiswa")
                    continue
                
                email = f"{nim}@student.ac.id"
                if User.objects.filter(email=email).exists():
                    errors.append(f"Baris {row_num}: Email '{email}' sudah digunakan")
                    continue
                
                try:
                    user = User.objects.create_user(
                        username=nim,
                        password=nim,
                        first_name=nama,
                        email=email,
                        role='MAHASISWA',
                        is_password_changed=False,
                        is_active=True
                    )
                    
                    mahasiswa = Mahasiswa.objects.create(
                        user=user,
                        nim=nim,
                        kelas=kelas,
                        angkatan=angkatan,
                        status='AKTIF',
                    )
                    
                    imported_data.append({
                        'nim': nim,
                        'nama': nama,
                        'kelas': kelas,
                        'angkatan': angkatan,
                        'email': email,
                        'password': nim
                    })
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Baris {row_num}: Gagal menyimpan data - {str(e)}")
                    continue
        
        response_data = {
            'success': True,
            'message': f'Berhasil mengupload {success_count} data mahasiswa',
            'data': {
                'success_count': success_count,
                'data': imported_data,
                'errors': errors,
                'total_rows': len(df)
            }
        }
        
        if errors:
            response_data['warning'] = f'Ada {len(errors)} data yang gagal diimport'
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}'
        }, status=500)


@login_required
def download_template_mahasiswa(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Mahasiswa"
    
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['NIM', 'Nama', 'Kelas', 'Angkatan']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    sample_data = [
        ['2201010001', 'John Doe', 'A', '2022'],
        ['2201010002', 'Jane Smith', 'B', '2022'],
        ['2201010003', 'Budi Santoso', 'C', '2023'],
        ['', '', '', ''],
        ['', '', '', ''],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col_idx in [1, 3, 4] else 'left')
    
    for col in range(1, 5):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20
    
    ws['F1'] = '📝 Catatan:'
    ws['F2'] = '1. NIM harus unik'
    ws['F3'] = '2. Semua kolom wajib diisi'
    ws['F4'] = '3. Password default = NIM'
    ws['F5'] = '4. Format: .xlsx atau .xls'
    
    for row in range(1, 6):
        ws[f'F{row}'].font = Font(size=10, color='666666')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="template_mahasiswa.xlsx"'
    wb.save(response)
    
    return response


# =========================================================
# KEMBALIKAN SEMUA MAHASISWA DARI ARSIP
# =========================================================
@login_required
@transaction.atomic
def kembalikan_semua_arsip(request):
    has_access, response = check_role(request, ['KAPRODI'])
    if not has_access:
        return response

    if request.method == 'POST':
        ids = request.POST.getlist('mahasiswa_ids')
        if not ids:
            messages.warning(request, 'Pilih minimal satu mahasiswa.')
            return redirect('capstone_system:arsip_mahasiswa')

        mahasiswa_list = Mahasiswa.objects.filter(id__in=ids, status="ARSIP").select_related('user', 'dosen_pembimbing')
        
        if not mahasiswa_list.exists():
            messages.warning(request, 'Tidak ada mahasiswa arsip yang dipilih.')
            return redirect('capstone_system:arsip_mahasiswa')

        success_count = 0
        failed_count = 0
        failed_names = []

        dospem_to_update = set()

        for mahasiswa in mahasiswa_list:
            if mahasiswa.dosen_pembimbing:
                dospem = mahasiswa.dosen_pembimbing
                if dospem.penuh:
                    failed_count += 1
                    failed_names.append(mahasiswa.user.get_full_name())
                    continue
                dospem_to_update.add(dospem)

            mahasiswa.status = "AKTIF"
            mahasiswa.save()
            mahasiswa.user.is_active = True
            mahasiswa.user.save()
            success_count += 1

        for dospem in dospem_to_update:
            dospem.update_status()

        if success_count > 0:
            messages.success(
                request, 
                f'{success_count} mahasiswa berhasil dikembalikan dari arsip.'
            )
        
        if failed_count > 0:
            messages.error(
                request, 
                f'{failed_count} mahasiswa gagal dikembalikan karena kuota dosen penuh: {", ".join(failed_names)}'
            )

    return redirect('capstone_system:arsip_mahasiswa')